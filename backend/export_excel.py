"""
export_excel.py
----------------
Gera o arquivo de exportação já formatado, usando
config/templates/TIM_MW_SP_Preliminary_Report_-_Template.xlsx como molde.

Por que um template em vez de código:
    A formatação do controle (faixa azul-marinho no bloco de identificação,
    roxo nos END ID PTB, vermelho nas coordenadas, cores de tema no bloco de
    cronograma, larguras, painéis congelados, agrupamento de colunas e
    filtro) tem dezenas de regras. Reproduzir isso em Python daria um
    arquivo enorme e frágil.

    Mesma ideia já usada no CSV template: o padrão mora num arquivo que se
    edita no Excel, não em código. Para mudar uma cor, largura ou
    agrupamento do arquivo exportado, abra o template no Excel, ajuste a
    linha 1 e salve. Nada aqui precisa mudar.

    O template tem só duas linhas: a linha 1 (títulos, formatados) e a
    linha 2 (faixa de separação do original). Os dados entram a partir da
    linha 3.

O que este módulo faz é só: abrir o molde, despejar os registros e
carimbar o estilo das células de dados.
"""

import datetime as dt
import io
import os
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_XLSX = os.path.join(
    BASE_DIR, "config", "templates", "TIM_MW_SP_Preliminary_Report_-_Template.xlsx"
)

# Linha 1 = títulos, linha 2 = faixa de separação (vem do controle original).
DATA_START_ROW = 3

# Estilo das células de dados, igual ao do controle original: conteúdo
# centralizado com uma linha fina embaixo separando os registros.
DATA_FONT = Font(name="Calibri", size=11)
DATA_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DATA_BORDER = Border(bottom=Side(style="thin"))
DATE_FORMAT = "dd/mm/yyyy"
TEXT_FORMAT = "@"

# Regra do processo: TIM KEY sempre em xxxx.xxxxxx. A base tem valores
# quebrados porque o Excel trata a chave como número e descarta zeros à
# DIREITA ("2025.000610" vira "2025.00061"). Só zeros podem ter sido
# perdidos, então completar com zeros à direita reconstrói o valor exato.
# A DU Virtual segue o mesmo padrão, preservando o sufixo .LD_x criado
# a cada LOS Block. As duas células saem com formato TEXTO ("@") para o
# Excel nunca mais corromper esses valores.
_KEY_PATTERN = re.compile(r"^(\d{4})(?:[.,](\d{1,6}))?$")
_LD_SUFFIX = re.compile(r"^(.*?)\.(LD_\d+)$", re.IGNORECASE)
KEY_FIELDS = {"tim_key", "du_id_virtual"}


def format_tim_key(raw):
    text = str(raw or "").strip()
    match = _KEY_PATTERN.match(text)
    if not match:
        return text
    year, fraction = match.group(1), match.group(2) or ""
    return f"{year}.{fraction.ljust(6, '0')}"


def format_du_virtual(raw):
    text = str(raw or "").strip()
    if not text:
        return text
    match = _LD_SUFFIX.match(text)
    if match:
        return f"{format_tim_key(match.group(1))}.{match.group(2).upper()}"
    return format_tim_key(text)


class ExportError(Exception):
    """Levantada quando o template de formatação não pôde ser usado."""


def _cell_value(record, field):
    """
    Converte o valor guardado no banco para o que vai na célula.

    Datas viram date de verdade (para o Excel ordenar e filtrar como data,
    não como texto). Números viram número. O resto vai como está.
    """
    value = record.get(field["internal_name"])
    if value is None or value == "":
        return None

    if field["internal_name"] in KEY_FIELDS:
        return (format_tim_key if field["internal_name"] == "tim_key" else format_du_virtual)(value)

    if field["type"] == "date" and isinstance(value, str):
        try:
            return dt.datetime.strptime(value.strip(), "%Y-%m-%d").date()
        except ValueError:
            # Data em formato inesperado: melhor exportar o texto cru do que
            # descartar uma informação que alguém digitou à mão.
            return value

    if field["type"] in ("float", "integer") and isinstance(value, str):
        try:
            return float(value.replace(",", ".")) if field["type"] == "float" else int(value)
        except ValueError:
            return value

    return value


def build_workbook(records, schema):
    """
    Devolve um BytesIO com o .xlsx pronto (formatado + preenchido).

    records: lista de dicts (ver database.fetch_all_records)
    schema:  schema de campos - dita a ordem das colunas, igual ao CSV
    """
    if not os.path.exists(TEMPLATE_XLSX):
        raise ExportError(
            f"template de formatação não encontrado em {TEMPLATE_XLSX} - "
            "sem ele não é possível exportar com o padrão do controle."
        )

    ordered_fields = sorted(schema, key=lambda f: f["index"])

    workbook = load_workbook(TEMPLATE_XLSX)
    worksheet = workbook.active

    for row_offset, record in enumerate(records):
        row = DATA_START_ROW + row_offset
        for field in ordered_fields:
            cell = worksheet.cell(row=row, column=field["index"] + 1)
            cell.value = _cell_value(record, field)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if field["type"] == "date":
                cell.number_format = DATE_FORMAT
            elif field["internal_name"] in KEY_FIELDS:
                cell.number_format = TEXT_FORMAT

    # O filtro do template cobre só a linha de título; estende para os dados.
    last_column = get_column_letter(len(ordered_fields))
    last_row = DATA_START_ROW + len(records) - 1 if records else 1
    worksheet.auto_filter.ref = f"A1:{last_column}{max(last_row, 1)}"

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)
    return stream


def build_file_name(now=None):
    """TIM_MW_SP_Preliminary_Report_2026-07-22_1432.xlsx"""
    now = now or dt.datetime.now()
    return f"TIM_MW_SP_Preliminary_Report_{now:%Y-%m-%d_%H%M}.xlsx"
