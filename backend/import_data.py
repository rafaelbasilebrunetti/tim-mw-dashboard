"""
import_data.py
---------------
Importa em massa um controle existente (Excel .xlsx) para o banco do
dashboard, casando as colunas do arquivo com o schema gerado a partir
do CSV template (ver schema_loader.py).

USO:
    python import_data.py "caminho\\para\\seu_controle.xlsx"

    Opções:
    --sheet NOME_DA_ABA     Nome da aba a importar (padrão: primeira aba)
    --dry-run               Só mostra o que seria importado, não grava nada
    --upsert                Se já existir uma linha com o mesmo TIM KEY,
                             atualiza em vez de duplicar

O que o script faz:
    1. Lê a primeira linha do Excel como cabeçalho.
    2. Casa cada cabeçalho com um campo do schema (comparação tolerante:
       ignora maiúsculas/minúsculas, espaços extras e acentos).
    3. Se o arquivo tem o mesmo número de colunas do template, casa por
       POSIÇÃO (resolve a ambiguidade das colunas "SITE TYPE A/B"
       duplicadas do mesmo jeito que o schema_loader).
       Caso contrário, casa por NOME (aí colunas duplicadas de nome viram
       um alerta manual, pois não dá pra saber qual é qual só pelo nome).
    4. Converte datas do Excel (seriais ou datetime) para "YYYY-MM-DD".
    5. Insere no banco. Linhas totalmente vazias são puladas.
    6. Ao final, imprime um relatório: quantas linhas importadas, quais
       colunas do arquivo não foram reconhecidas, quais colunas do
       schema ficaram sem dado nenhuma linha.
"""

import argparse
import datetime
import os
import re
import sys
import unicodedata

from openpyxl import load_workbook

from database import TABLE_NAME, get_connection
from schema_loader import MANUAL_OVERRIDES, load_schema

# Consoles do Windows (cmd/PowerShell/Git Bash) costumam usar cp1252, que não
# tem os caracteres ⚠/ℹ/✓/✗ usados no relatório abaixo e derruba o script com
# UnicodeEncodeError bem no fim da importação. Troca por "?" em vez de travar.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(errors="replace")


def normalize(text: str) -> str:
    if text is None:
        return ""
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def excel_value_to_str(value, field_type: str):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        # Se a célula tem só hora 00:00:00 sobrando de formatação do Excel,
        # ainda assim é uma data válida - mantém.
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.time):
        # Célula formatada como HORA (não data) por engano na planilha
        # original. Não existe "data" real aqui - deixa vazio em vez de
        # gravar um texto sem sentido; isso aparece no relatório final.
        return None
    if field_type == "float":
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Planilhas brasileiras costumam usar vírgula decimal
            # (ex: "-23,9308"). Converte para ponto antes de gravar,
            # senão o banco guarda como texto e a API quebra depois.
            cleaned = value.strip().replace(",", ".")
            try:
                return float(cleaned)
            except ValueError:
                return None  # não deu pra converter - fica vazio em vez de quebrar
        return None
    if field_type == "date" and isinstance(value, str):
        return value.strip() or None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def build_column_mapping(file_headers, schema):
    """
    Retorna:
      - mapping: lista paralela a file_headers, cada item é o field do
        schema (dict) que essa coluna do arquivo alimenta, ou None se
        não foi possível casar.
      - unmatched_schema: campos do schema que nenhuma coluna do
        arquivo alimentou.
    """
    mapping = [None] * len(file_headers)

    if len(file_headers) == len(schema):
        # Mesmo número de colunas do template -> assume mesma ordem
        # (mesma lógica de posição usada no schema_loader para resolver
        # as colunas "SITE TYPE A/B" duplicadas).
        mismatches = []
        for i, (file_header, field) in enumerate(zip(file_headers, schema)):
            mapping[i] = field
            if normalize(file_header) != normalize(field["csv_header"]):
                mismatches.append((i, file_header, field["csv_header"]))
        if mismatches:
            print(f"\n⚠ {len(mismatches)} cabeçalho(s) com texto diferente do template "
                  f"(mesma posição foi assumida mesmo assim):")
            for i, got, expected in mismatches[:15]:
                print(f'   coluna {i}: arquivo="{got}"  template="{expected}"')
    else:
        # Número de colunas diferente -> casa por nome normalizado.
        # Constrói índice nome_normalizado -> [fields] (pode ter mais de
        # um, ex: "SITE TYPE A" aparece 2x no template).
        by_name = {}
        for field in schema:
            by_name.setdefault(normalize(field["csv_header"]), []).append(field)

        for i, file_header in enumerate(file_headers):
            candidates = by_name.get(normalize(file_header))
            if not candidates:
                continue
            # usa o primeiro candidato ainda não usado
            for field in candidates:
                if field not in mapping:
                    mapping[i] = field
                    break

    used_fields = {f["internal_name"] for f in mapping if f}
    unmatched_schema = [f for f in schema if f["internal_name"] not in used_fields]
    return mapping, unmatched_schema


def read_rows_from_workbook(wb, sheet_name=None):
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    file_headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    data_rows = [row for row in rows_iter]
    return file_headers, data_rows


def read_rows(path, sheet_name=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    file_headers, data_rows = read_rows_from_workbook(wb, sheet_name)
    wb.close()
    return file_headers, data_rows


def process_rows(file_headers, data_rows, schema, conn, dry_run=False, upsert=False):
    """
    Núcleo do import, sem I/O de arquivo nem prints - usado tanto pelo
    CLI (import_file) quanto pelo endpoint web (import_routes.py).
    Retorna um relatório estruturado (dict) em vez de imprimir.
    """
    mapping, unmatched_schema = build_column_mapping(file_headers, schema)
    unmatched_file_cols = [h for h, f in zip(file_headers, mapping) if f is None and h]

    imported, skipped_empty, updated, failed = 0, 0, 0, []
    suspicious_time_cells = []

    for row_num, row in enumerate(data_rows, start=2):  # linha 2 = primeira linha de dados (após cabeçalho)
        record = {}
        for value, field in zip(row, mapping):
            if field is None:
                continue
            converted = excel_value_to_str(value, field["type"])
            record[field["internal_name"]] = converted
            if field["type"] == "date" and isinstance(value, datetime.time):
                suspicious_time_cells.append((row_num, field["csv_header"], value))

        if not any(v not in (None, "") for v in record.values()):
            skipped_empty += 1
            continue

        if dry_run:
            imported += 1
            continue

        tim_key = record.get("tim_key")
        existing_id = None
        try:
            if upsert and tim_key:
                existing = conn.execute(
                    f'SELECT id FROM {TABLE_NAME} WHERE "tim_key" = ?', (tim_key,)
                ).fetchone()
                if existing:
                    existing_id = existing["id"]

            if existing_id:
                set_clause = ", ".join(f'"{k}" = ?' for k in record.keys())
                conn.execute(
                    f'UPDATE {TABLE_NAME} SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                    list(record.values()) + [existing_id],
                )
                updated += 1
            else:
                columns = ", ".join(f'"{k}"' for k in record.keys())
                placeholders = ", ".join("?" for _ in record)
                conn.execute(
                    f'INSERT INTO {TABLE_NAME} ({columns}) VALUES ({placeholders})',
                    list(record.values()),
                )
                imported += 1
        except Exception as exc:
            # Uma linha ruim não deve derrubar a importação inteira -
            # registra o erro e segue para a próxima.
            failed.append({"row": row_num, "tim_key": tim_key, "error": str(exc)})

    if not dry_run:
        conn.commit()

    return {
        "imported": imported,
        "updated": updated,
        "skipped_empty": skipped_empty,
        "failed": failed,
        "unmatched_file_columns": unmatched_file_cols,
        "unmatched_schema_fields": [f["csv_header"] for f in unmatched_schema],
        "suspicious_time_cells": [
            {"row": r, "column": h, "value": str(v)} for r, h, v in suspicious_time_cells
        ],
    }


def import_file(path, sheet_name=None, dry_run=False, upsert=False):
    if not os.path.exists(path):
        print(f"Arquivo não encontrado: {path}")
        sys.exit(1)

    schema = load_schema()
    file_headers, data_rows = read_rows(path, sheet_name)

    print(f"Arquivo: {path}")
    print(f"Colunas encontradas: {len(file_headers)}  |  Colunas no schema: {len(schema)}")
    print(f"Linhas de dados encontradas: {len(data_rows)}")

    conn = None if dry_run else get_connection()
    report = process_rows(file_headers, data_rows, schema, conn, dry_run=dry_run, upsert=upsert)
    if not dry_run:
        conn.close()

    if report["unmatched_file_columns"]:
        print(f"\n⚠ {len(report['unmatched_file_columns'])} coluna(s) do arquivo não reconhecida(s) "
              f"(serão ignoradas): {report['unmatched_file_columns']}")
    if report["unmatched_schema_fields"]:
        print(f"\nℹ {len(report['unmatched_schema_fields'])} campo(s) do schema sem coluna correspondente "
              f"no arquivo (ficarão vazios): {report['unmatched_schema_fields']}")

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Concluído:")
    print(f"  Novas linhas importadas: {report['imported']}")
    if upsert:
        print(f"  Linhas atualizadas (TIM KEY já existia): {report['updated']}")
    print(f"  Linhas vazias puladas: {report['skipped_empty']}")

    if report["suspicious_time_cells"]:
        print(f"\n⚠ {len(report['suspicious_time_cells'])} célula(s) com valor de HORA em campo de DATA "
              f"(provável erro de formatação na planilha original - ficaram vazias no banco, revise):")
        for cell in report["suspicious_time_cells"][:15]:
            print(f'   linha {cell["row"]}, coluna "{cell["column"]}": {cell["value"]}')

    if report["failed"]:
        print(f"\n✗ {len(report['failed'])} linha(s) falharam e foram puladas:")
        for item in report["failed"][:15]:
            print(f'   linha {item["row"]} (TIM KEY={item["tim_key"]}): {item["error"]}')


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importa um controle Excel existente para o dashboard.")
    parser.add_argument("file", help="Caminho para o arquivo .xlsx")
    parser.add_argument("--sheet", default=None, help="Nome da aba (padrão: primeira aba)")
    parser.add_argument("--dry-run", action="store_true", help="Não grava no banco, só mostra o que seria feito")
    parser.add_argument("--upsert", action="store_true", help="Atualiza linhas existentes com o mesmo TIM KEY em vez de duplicar")
    args = parser.parse_args()

    import_file(args.file, sheet_name=args.sheet, dry_run=args.dry_run, upsert=args.upsert)
