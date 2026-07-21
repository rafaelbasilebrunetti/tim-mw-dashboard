"""
site_reference.py
------------------
Enriquece o detalhe de um site com informações que não vêm do CSV
template do dashboard (Infra Type, Município, Detentora, Lat/Long),
buscando numa planilha de referência externa e maior (cadastro de
sites da rede), mantida fora do controle do dashboard.

Onde colocar o arquivo:
    data/site_reference.xlsb   (preferencial - Excel binário)
    data/site_reference.xlsx   (alternativa, se convertido)

Formato esperado (colunas por LETRA - a planilha real não tem nomes de
coluna padronizados, então o mapeamento é posicional):
    A  (0)  - Site ID (chave de busca)
    E  (4)  - Endereço ID (END ID)
    H  (7)  - Detentor da área (Detentora)
    R  (17) - Cidade (Município)
    AB (27) - Tipo de Infra (Infra Type)
    AC (28) - Latitude
    AD (29) - Longitude
    AA (26) - Data última alteração (usada só como desempate, ver abaixo)

A primeira linha é tratada como cabeçalho e ignorada.

Por que não usar o Endereço ID (END ID) como uma segunda chave de
busca: confirmado nos dados reais que o mesmo Endereço ID aparece em
várias linhas com Site ID diferente (endereço/torre compartilhado por
vários elementos) - buscar por END ID poderia trazer os dados de outro
site. Por isso a busca é sempre pela própria linha do Site ID.

Duplicatas: ~3% dos Site IDs aparecem em mais de uma linha com valores
diferentes entre si. Quando isso acontece, fica a linha com a "Data
última alteração" (coluna AA) mais recente.

Cache: o arquivo (pode ter 300 mil+ linhas) é lido uma única vez e
mantido em memória; só é recarregado se o mtime do arquivo mudar. A
primeira consulta após o backend subir pode demorar dezenas de
segundos - as seguintes são instantâneas (dict em memória).
"""

import os
import threading

from openpyxl import load_workbook
from pyxlsb import open_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
REFERENCE_XLSB = os.path.join(DATA_DIR, "site_reference.xlsb")
REFERENCE_XLSX = os.path.join(DATA_DIR, "site_reference.xlsx")

COL_SITE_ID = 0
COL_END_ID = 4
COL_DETENTORA = 7
COL_MUNICIPIO = 17
COL_INFRA_TYPE = 27
COL_LAT = 28
COL_LONG = 29
COL_LAST_CHANGE = 26

_lock = threading.Lock()
_cache = {"path": None, "mtime": None, "by_site_id": {}}


def _active_reference_path():
    if os.path.exists(REFERENCE_XLSB):
        return REFERENCE_XLSB
    if os.path.exists(REFERENCE_XLSX):
        return REFERENCE_XLSX
    return None


def _normalize(text):
    return str(text).strip().upper() if text not in (None, "") else ""


def _cell(row, index):
    value = row.get(index)
    return value if value not in (None, "") else None


def _iter_rows_xlsb(path):
    with open_workbook(path) as wb:
        with wb.get_sheet(1) as sheet:
            for i, row in enumerate(sheet.rows()):
                if i == 0 or not row:
                    continue
                yield {cell.c: cell.v for cell in row}


def _iter_rows_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row:
                continue
            yield {idx: value for idx, value in enumerate(row)}
    finally:
        wb.close()


def _build_index(path):
    iter_rows = _iter_rows_xlsb if path.lower().endswith(".xlsb") else _iter_rows_xlsx
    by_site_id = {}

    for row in iter_rows(path):
        site_id = _cell(row, COL_SITE_ID)
        if not site_id:
            continue
        key = _normalize(site_id)

        last_change = str(row.get(COL_LAST_CHANGE) or "")
        existing = by_site_id.get(key)
        if existing is not None and existing["_last_change"] >= last_change:
            continue

        by_site_id[key] = {
            "_last_change": last_change,
            "end_id": _cell(row, COL_END_ID),
            "detentora": _cell(row, COL_DETENTORA),
            "municipio": _cell(row, COL_MUNICIPIO),
            "infra_type": _cell(row, COL_INFRA_TYPE),
            "lat": _cell(row, COL_LAT),
            "long": _cell(row, COL_LONG),
        }

    return by_site_id


def _get_index():
    path = _active_reference_path()
    if path is None:
        return None

    mtime = os.path.getmtime(path)
    if _cache["path"] == path and _cache["mtime"] == mtime:
        return _cache["by_site_id"]

    with _lock:
        # outra thread pode ter terminado de recarregar enquanto esperávamos o lock
        if _cache["path"] == path and _cache["mtime"] == mtime:
            return _cache["by_site_id"]
        by_site_id = _build_index(path)
        _cache.update({"path": path, "mtime": mtime, "by_site_id": by_site_id})
        return by_site_id


def lookup_site(site_name):
    """
    Retorna as informações de referência de um site pelo nome (Site A
    ou Site B). Nunca levanta exceção por site não encontrado - sempre
    devolve um dict com "found" indicando o resultado.
    """
    if not site_name:
        return {"found": False}

    index = _get_index()
    if index is None:
        return {"found": False, "reference_missing": True}

    row = index.get(_normalize(site_name))
    if row is None:
        return {"found": False}

    return {
        "found": True,
        "end_id": row["end_id"],
        "infra_type": row["infra_type"],
        "municipio": row["municipio"],
        "detentora": row["detentora"],
        "lat": row["lat"],
        "long": row["long"],
    }
